import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

class ExcelInfoSaver:
    def __init__(self, filename, company, info):
        self.filename = filename
        self.company = company
        self.info = info
        self.wb = self._load_or_create_workbook()
        self.ws = self._create_or_get_sheet()
        self.bold_font = Font(bold=True)
    
    def _load_or_create_workbook(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        else:
            wb = load_workbook(self.filename)
        return wb
    
    def _create_or_get_sheet(self):
        return self.wb.create_sheet(title=self.company)
    
    def _add_headers(self, row, headers):
        for col_idx, header in enumerate(headers, start=1):
            self.ws.cell(row=row, column=col_idx, value=header).font = self.bold_font
    
    def _add_section_title(self, row, title):
        self.ws.cell(row=row, column=1, value=title).font = self.bold_font
    
    def _fill_data_rows(self, start_row, data):
        for i, row_data in enumerate(data, start=start_row):
            for j, cell_val in enumerate(row_data, start=1):
                self.ws.cell(row=i, column=j, value=cell_val)
    
    def _populate_general_info(self):
        self._add_section_title(1, "Información General")
        self._add_headers(2, ["Campo", "Descripción", "Valor", "Fuente de Información"])
        
        data_info = [
            ["Nombre Legal", "Nombre completo y legal de la empresa", "", ""],
            ["Sector Industrial", "Sector industrial donde opera la empresa", "", ""],
            ["Facturación Anual", "Facturación anual de la empresa", "", ""],
            ["Número de Empleados", "Número total de empleados", "", ""],
            ["Ubicación", "Ubicación geográfica principal", "", ""],
            ["Sitio Web Oficial", "URL del sitio web oficial", "", ""],
            ["Estructura Organizacional", "Tipo de estructura organizacional", "", ""]
        ]
        self._fill_data_rows(3, data_info)
        
        info_fields = {
            "Nombre Legal": ("nombre_legal", 3),
            "Sector Industrial": ("sector_industrial", 4),
            "Facturación Anual": ("facturacion_anual", 5),
            "Número de Empleados": ("numero_empleados", 6),
            "Ubicación": ("ubicacion", 7),
            "Sitio Web Oficial": ("sitio_web_oficial", 8),
            "Estructura Organizacional": ("estructura_organizacional", 9)
        }
        
        general_info = self.info.get("informacion_general", {})
        for _, (field_key, row_num) in info_fields.items():
            field_data = general_info.get(field_key, {})
            value = field_data.get("valor", "")
            source = "; ".join(field_data.get("fuente", [])) if isinstance(field_data.get("fuente"), list) else field_data.get("fuente", "")
            self.ws.cell(row=row_num, column=3, value=value)
            self.ws.cell(row=row_num, column=4, value=source)
    
    def _populate_decision_makers(self, start_row=12):
        self._add_section_title(start_row, "Tomadores de Decisión")
        self._add_headers(start_row + 1, ["Cargo", "Nombre", "Fuente de Información"])
        
        tomadores = self.info.get("tomadores_decision", [])
        for i, item in enumerate(tomadores, start=start_row + 2):
            self.ws.cell(row=i, column=1, value=item.get("rol", ""))
            self.ws.cell(row=i, column=2, value=item.get("nombre", ""))
            self.ws.cell(row=i, column=3, value=item.get("fuente", ""))
        return len(tomadores)
    
    def _populate_strategy_context(self, start_row):
        self._add_section_title(start_row + 2, "Estrategia y Contexto")
        self._add_headers(start_row + 3, ["Campo", "Descripción", "Información Detallada", "Fuente de Información"])
        
        info_fields = {
            "Estrategia Tecnológica": "estrategia_tecnologica",
            "Contexto de Mercado": "contexto_mercado",
            "Objetivos del CEO": "objetivos_ceo"
        }
        
        descriptions = {
            "Estrategia Tecnológica": "Tecnologías que la empresa planea adoptar y objetivos que pretende conseguir",
            "Contexto de Mercado": "Contexto actual del mercado y cómo afecta a la estrategia de la empresa",
            "Objetivos del CEO": "Objetivos futuros mencionados por el CEO"
        }
        
        estrategia = self.info.get("estrategia_y_contexto", {})
        for i, (campo, clave) in enumerate(info_fields.items(), start=start_row + 4):
            self.ws.cell(row=i, column=1, value=campo)
            self.ws.cell(row=i, column=2, value=descriptions.get(campo, ""))
            self.ws.cell(row=i, column=3, value=estrategia.get(clave, {}).get("valor", ""))
            fuente = estrategia.get(clave, {}).get("fuente", [])
            source_value = "; ".join(fuente) if isinstance(fuente, list) else fuente
            self.ws.cell(row=i, column=4, value=source_value)
    
    def _adjust_column_widths(self):
        for col in self.ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            self.ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    def save(self):
        self._populate_general_info()
        decision_maker_count = self._populate_decision_makers()
        strategy_start_row = 14 + decision_maker_count  # Adjusted dynamically
        self._populate_strategy_context(strategy_start_row)
        self._adjust_column_widths()
        self.wb.save(self.filename)
        return f"Información de {self.company} ha sido actualizada en {self.filename}"

# Usage Example
# info_saver = ExcelInfoSaver("output.xlsx", "NVIDIA", info_data)
# result_message = info_saver.save()
# print(result_message)
